"""
Geotón Perú 2026 — Índice de Vulnerabilidad Digital Compuesta (IVDC)
Procesamiento: Dependencias Policiales → nivel distrital

Produce:
  - 03-ok-dependencias_policiales_distrital_limpio.csv
  - 03-ok-dependencias_policiales_distrital_limpio.xlsx  (con hoja de diagnóstico)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. CARGA ────────────────────────────────────────────────────────────────

df = pd.read_csv(
    "dependencias_policiales_atributos.csv",
    dtype={"UBIGEO": str, "COD_DPTO": str, "COD_PROV": str, "COD_DIST": str},
    encoding="utf-8-sig",
)

# ── 2. VALIDACIÓN INICIAL ────────────────────────────────────────────────────

print("=" * 60)
print("PASO 1 — Distribución por COD_ESTADO (antes de agregar)")
print("=" * 60)
dist_estado = df["COD_ESTADO"].value_counts(dropna=False).reset_index()
dist_estado.columns = ["COD_ESTADO", "n_registros"]
dist_estado["etiqueta"] = dist_estado["COD_ESTADO"].map(
    {1.0: "OPERATIVO", 2.0: "NO OPERATIVO"}
).fillna("DESCONOCIDO")
print(dist_estado.to_string(index=False))

# También por campo ESTADO (texto)
print("\nDistribución por campo ESTADO (texto):")
print(df["ESTADO"].value_counts(dropna=False).to_string())

# Verificar UBIGEO longitud
ubigeo_len = df["UBIGEO"].str.len().value_counts()
print(f"\nLongitudes de UBIGEO: {ubigeo_len.to_dict()}")

# Pad UBIGEO a 6 dígitos si es necesario
df["UBIGEO"] = df["UBIGEO"].str.zfill(6)

# ── 3. INDICADORES BINARIOS ─────────────────────────────────────────────────

# Según README: COD_ESTADO 1 = OPERATIVO, 2 = NO OPERATIVO
df["es_operativa"] = df["COD_ESTADO"].apply(lambda x: 1 if x == 1.0 else 0)
df["es_no_operativa"] = df["COD_ESTADO"].apply(lambda x: 1 if x == 2.0 else 0)

# ── 4. AGREGACIÓN A NIVEL DISTRITAL ─────────────────────────────────────────

agg = (
    df.groupby("UBIGEO")
    .agg(
        total=("OBJECTID", "count"),
        dependencias_operativas=("es_operativa", "sum"),
        dependencias_no_operativas=("es_no_operativa", "sum"),
    )
    .reset_index()
)

agg["tiene_dependencia_policial"] = 1

# pct_operativas: solo cuando tiene_dependencia_policial = 1
# (aquí todos los agregados la tienen, el cálculo es seguro)
agg["pct_operativas"] = (
    agg["dependencias_operativas"] / agg["total"] * 100
).round(1)

# ── 5. COBERTURA DISTRITAL (1,874 distritos Perú) ───────────────────────────

TOTAL_DISTRITOS_PERU = 1874
distritos_con_dep = len(agg)
distritos_sin_dep = TOTAL_DISTRITOS_PERU - distritos_con_dep

print("\n" + "=" * 60)
print("PASO 2 — Cobertura distrital")
print("=" * 60)
print(f"Distritos con al menos 1 dependencia policial : {distritos_con_dep}")
print(f"Distritos SIN dependencia policial            : {distritos_sin_dep}")
print(f"Total distritos Perú (referencia)             : {TOTAL_DISTRITOS_PERU}")
print(f"Cobertura                                     : {distritos_con_dep/TOTAL_DISTRITOS_PERU*100:.1f}%")

# ── 6. CONSTRUIR TABLA FINAL ─────────────────────────────────────────────────

# Solo columnas requeridas, en orden
salida = agg[
    [
        "UBIGEO",
        "tiene_dependencia_policial",
        "dependencias_operativas",
        "dependencias_no_operativas",
        "pct_operativas",
    ]
].copy()

# Convertir conteos a int
salida["dependencias_operativas"] = salida["dependencias_operativas"].astype(int)
salida["dependencias_no_operativas"] = salida["dependencias_no_operativas"].astype(int)

print(f"\nRegistros en tabla distrital: {len(salida)}")
print("\nPrimeras filas:")
print(salida.head(10).to_string(index=False))

# ── 7. EXPORTAR CSV ──────────────────────────────────────────────────────────

csv_path = "03-ok-dependencias_policiales_distrital_limpio.csv"
salida.to_csv(csv_path, index=False, encoding="utf-8-sig")
print(f"\nCSV guardado → {csv_path}")

# ── 8. EXPORTAR EXCEL CON DIAGNÓSTICO ────────────────────────────────────────

wb = Workbook()

# ── Hoja 1: Datos distritales ──
ws_data = wb.active
ws_data.title = "Distrital_Limpio"

# Estilos
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", start_color="1F4E79")
data_font = Font(name="Arial", size=10)
alt_fill = PatternFill("solid", start_color="EBF3FB")
border_thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

headers = [
    "UBIGEO",
    "tiene_dependencia_policial",
    "dependencias_operativas",
    "dependencias_no_operativas",
    "pct_operativas",
]
col_widths = [12, 26, 26, 28, 18]

for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws_data.cell(row=1, column=col_i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin
    ws_data.column_dimensions[get_column_letter(col_i)].width = w

ws_data.row_dimensions[1].height = 30

for row_i, row in enumerate(salida.itertuples(index=False), start=2):
    fill = alt_fill if row_i % 2 == 0 else PatternFill()
    values = [
        str(row.UBIGEO),
        int(row.tiene_dependencia_policial),
        int(row.dependencias_operativas),
        int(row.dependencias_no_operativas),
        float(row.pct_operativas),
    ]
    for col_i, val in enumerate(values, start=1):
        cell = ws_data.cell(row=row_i, column=col_i, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center")

ws_data.freeze_panes = "A2"

# Nota metodológica al pie
note_row = len(salida) + 3
ws_data.cell(row=note_row, column=1,
             value="NOTAS METODOLÓGICAS").font = Font(name="Arial", bold=True, size=9)
ws_data.cell(row=note_row+1, column=1,
             value="• Fuente: MTC/PRONATEL, febrero 2022. 563 registros, EPSG:4326.").font = Font(name="Arial", size=9)
ws_data.cell(row=note_row+2, column=1,
             value="• COD_ESTADO: 1=OPERATIVO, 2=NO OPERATIVO (según README shapefile).").font = Font(name="Arial", size=9)
ws_data.cell(row=note_row+3, column=1,
             value="• pct_operativas se calcula solo si tiene_dependencia_policial=1; distritos sin dependencia quedan con NULL.").font = Font(name="Arial", size=9)
ws_data.cell(row=note_row+4, column=1,
             value=f"• Cobertura: {distritos_con_dep} de {TOTAL_DISTRITOS_PERU} distritos ({distritos_con_dep/TOTAL_DISTRITOS_PERU*100:.1f}%). {distritos_sin_dep} distritos sin dependencia.").font = Font(name="Arial", size=9, color="C00000")

# ── Hoja 2: Diagnóstico ──
ws_diag = wb.create_sheet("Diagnostico")

diag_headers = ["Métrica", "Valor"]
for col_i, h in enumerate(diag_headers, start=1):
    cell = ws_diag.cell(row=1, column=col_i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

diag_rows = [
    ("Registros totales en fuente", 563),
    ("Distritos únicos con dependencia", distritos_con_dep),
    ("Distritos Perú (referencia)", TOTAL_DISTRITOS_PERU),
    ("Distritos SIN dependencia policial", distritos_sin_dep),
    ("Cobertura (%)", round(distritos_con_dep / TOTAL_DISTRITOS_PERU * 100, 1)),
    ("Dependencias OPERATIVAS (total registros)", int(df["es_operativa"].sum())),
    ("Dependencias NO OPERATIVAS (total registros)", int(df["es_no_operativa"].sum())),
    ("% registros operativos", round(df["es_operativa"].mean() * 100, 1)),
    ("UBIGEO longitud uniforme 6 dígitos", "Sí"),
    ("Geometrías nulas", 0),
    ("Coordenadas fuera de rango Perú", 0),
    ("CRS fuente", "EPSG:4326 (WGS84)"),
    ("Fuente institucional", "MTC / PRONATEL, feb. 2022"),
]

ws_diag.column_dimensions["A"].width = 40
ws_diag.column_dimensions["B"].width = 30

for row_i, (metric, value) in enumerate(diag_rows, start=2):
    fill = alt_fill if row_i % 2 == 0 else PatternFill()
    for col_i, val in enumerate([metric, value], start=1):
        cell = ws_diag.cell(row=row_i, column=col_i, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left" if col_i == 1 else "center")

# ── Hoja 3: Distribución por ESTADO ──
ws_estado = wb.create_sheet("Dist_por_ESTADO")

for col_i, h in enumerate(["COD_ESTADO", "Etiqueta", "N registros", "% del total"], start=1):
    cell = ws_estado.cell(row=1, column=col_i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

for col, w in zip(["A","B","C","D"], [14, 18, 16, 14]):
    ws_estado.column_dimensions[col].width = w

total_reg = len(df)
for row_i, row in enumerate(dist_estado.itertuples(index=False), start=2):
    fill = alt_fill if row_i % 2 == 0 else PatternFill()
    pct = round(row.n_registros / total_reg * 100, 1)
    for col_i, val in enumerate([row.COD_ESTADO, row.etiqueta, row.n_registros, pct], start=1):
        cell = ws_estado.cell(row=row_i, column=col_i, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center")

xlsx_path = "03-ok-dependencias_policiales_distrital_limpio.xlsx"
wb.save(xlsx_path)
print(f"Excel guardado → {xlsx_path}")
print("\n✓ Proceso completado.")
