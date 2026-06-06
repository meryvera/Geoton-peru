"""
Geotón Perú 2026 — IVDC
Genera Excel con dos hojas agregadas a nivel distrital:
  1. escuelas_distrital
  2. salud_distrital
Más una hoja de log con el resumen de limpieza.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
ESC_PATH  = "/content/locales_escolares_atributos.csv"
SAL_PATH  = "/content/salud_pronatel_atributos.csv"
OUT_XLSX  = os.path.join(BASE, "salud_y_escuelas.xlsx")

log_lines = []

# ── 1. LEER ──────────────────────────────────────────────────────────────────
esc_raw = pd.read_csv(ESC_PATH, dtype=str, encoding="utf-8-sig")
sal_raw = pd.read_csv(SAL_PATH, dtype=str, encoding="utf-8-sig")

log_lines.append(f"Escuelas leídas: {len(esc_raw)} filas, {esc_raw.shape[1]} columnas")
log_lines.append(f"Salud leída:     {len(sal_raw)} filas, {sal_raw.shape[1]} columnas")

# ── 2. VERIFICAR VALORES ÚNICOS DE ESTADO ─────────────────────────────────
esc_estados = esc_raw["ESTADO"].value_counts().to_dict()
sal_estados = sal_raw["ESTADO"].value_counts().to_dict()
log_lines.append(f"Escuelas ESTADO valores únicos: {esc_estados}")
log_lines.append(f"Salud    ESTADO valores únicos: {sal_estados}")

# ── 3. NORMALIZAR UBIGEO ──────────────────────────────────────────────────
# Escuelas: UBIGEO = COD_DIST (6 dígitos)
esc_raw["UBIGEO"] = esc_raw["COD_DIST"].str.strip().str.zfill(6)

# Salud: columna UBIGEO ya existe
sal_raw["UBIGEO"] = sal_raw["UBIGEO"].str.strip().str.zfill(6)

# Validar longitud
esc_bad_ubigeo = (esc_raw["UBIGEO"].str.len() != 6).sum()
sal_bad_ubigeo = (sal_raw["UBIGEO"].str.len() != 6).sum()
log_lines.append(f"Escuelas UBIGEO con longitud ≠ 6: {esc_bad_ubigeo}")
log_lines.append(f"Salud    UBIGEO con longitud ≠ 6: {sal_bad_ubigeo}")

# ── 4. DEDUPLICAR ESCUELAS ────────────────────────────────────────────────
# README: 83 filas con misma combinación local+CCPP+territorio
dup_cols_esc = ["NOM_LOC_ES", "COD_CCPP", "UBIGEO"]
esc_antes = len(esc_raw)
esc_clean = esc_raw.drop_duplicates(subset=dup_cols_esc, keep="first")
esc_eliminados = esc_antes - len(esc_clean)
log_lines.append(f"Escuelas duplicados eliminados (NOM_LOC_ES+COD_CCPP+UBIGEO): {esc_eliminados}")
log_lines.append(f"Escuelas después de deduplicar: {len(esc_clean)}")

# Salud: sin duplicados según README (no se aplica dedup)
sal_clean = sal_raw.copy()
log_lines.append(f"Salud sin deduplicación aplicada (README no reporta duplicados exactos)")

# ── 5. CLASIFICAR ESTADO ─────────────────────────────────────────────────
# Valores confirmados: "OPERATIVO" / "NO OPERATIVO"
esc_clean["es_operativo"] = esc_clean["ESTADO"].str.strip().str.upper() == "OPERATIVO"
sal_clean["es_operativo"] = sal_clean["ESTADO"].str.strip().str.upper() == "OPERATIVO"

# ── 6. AGREGAR ESCUELAS ───────────────────────────────────────────────────
esc_agg = (
    esc_clean.groupby("UBIGEO")
    .agg(
        total_escuelas_pronatel=("UBIGEO", "count"),
        escuelas_operativas=("es_operativo", "sum"),
        escuelas_no_operativas=("es_operativo", lambda x: (~x).sum()),
    )
    .reset_index()
)
esc_agg["pct_escuelas_operativas"] = (
    esc_agg["escuelas_operativas"] / esc_agg["total_escuelas_pronatel"]
).round(4)
esc_agg = esc_agg.sort_values("UBIGEO").reset_index(drop=True)

# ── 7. AGREGAR SALUD ──────────────────────────────────────────────────────
sal_agg = (
    sal_clean.groupby("UBIGEO")
    .agg(
        total_establecimientos_salud_pronatel=("UBIGEO", "count"),
        establecimientos_operativos=("es_operativo", "sum"),
        establecimientos_no_operativos=("es_operativo", lambda x: (~x).sum()),
    )
    .reset_index()
)
sal_agg["pct_establecimientos_operativos"] = (
    sal_agg["establecimientos_operativos"]
    / sal_agg["total_establecimientos_salud_pronatel"]
).round(4)
sal_agg = sal_agg.sort_values("UBIGEO").reset_index(drop=True)

log_lines.append(f"Distritos con escuelas PRONATEL: {len(esc_agg)}")
log_lines.append(f"Distritos con salud PRONATEL:    {len(sal_agg)}")
log_lines.append(f"Total escuelas (post-dedup):      {esc_agg['total_escuelas_pronatel'].sum()}")
log_lines.append(f"Total establec. salud:            {sal_agg['total_establecimientos_salud_pronatel'].sum()}")

# ── 8. CONSTRUIR EXCEL ────────────────────────────────────────────────────
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    esc_agg.to_excel(writer, sheet_name="escuelas_distrital", index=False)
    sal_agg.to_excel(writer, sheet_name="salud_distrital", index=False)
    
    # Hoja de log
    log_df = pd.DataFrame({"log": log_lines})
    log_df.to_excel(writer, sheet_name="log_limpieza", index=False)

# ── 9. FORMATEAR ─────────────────────────────────────────────────────────
wb = load_workbook(OUT_XLSX)

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center")
LEFT        = Alignment(horizontal="left")
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALT_FILL = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")

def format_sheet(ws, pct_col_idx):
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, cell in enumerate(row, start=1):
            cell.font = DATA_FONT
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if col_idx == 1:  # UBIGEO → texto, izquierda
                cell.alignment = LEFT
                cell.number_format = "@"
            elif col_idx == pct_col_idx:  # porcentaje
                cell.number_format = "0.00%"
                cell.alignment = CENTER
            else:
                cell.alignment = CENTER

    # Auto-ancho
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

format_sheet(wb["escuelas_distrital"], pct_col_idx=5)
format_sheet(wb["salud_distrital"],    pct_col_idx=5)

# Log sheet simple
ws_log = wb["log_limpieza"]
ws_log.column_dimensions["A"].width = 80
for cell in ws_log["A"]:
    cell.font = DATA_FONT

wb.save(OUT_XLSX)
print(f"✅ Archivo generado: {OUT_XLSX}")
for line in log_lines:
    print(" ", line)
