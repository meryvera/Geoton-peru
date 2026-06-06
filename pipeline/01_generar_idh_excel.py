"""
Geotón Perú 2026 — Índice de Vulnerabilidad Digital Compuesta (IVDC)
Preparación y análisis exploratorio del IDH distrital del Perú

Dataset requerido : peru_idh_distrital_atributos.csv  (PNUD 2019)
Output principal  : idh_distrital_limpio.xlsx  (varias hojas)
Output secundario : idh_distrital_limpio.csv

CONVENCIÓN DE QUINTILES (diferente al dataset original):
    quintil_idh = 1  →  IDH más BAJO  (cuartil inferior)
    quintil_idh = 5  →  IDH más ALTO  (cuartil superior)
    El dataset original usaba 1 = más alto; aquí se recalcula.

ADVERTENCIA UBIGEO:
    Si abres el CSV con doble clic en Excel, los ceros iniciales del UBIGEO
    pueden perderse (ej. "010101" → 10101).  Para evitarlo, importa el CSV
    indicando la columna UBIGEO como texto (Formato de datos: Texto).
    En el Excel generado aquí, UBIGEO ya está protegido con formato @.

Uso:
    python generar_idh_excel.py [ruta_csv]

    Si no se pasa ruta, busca peru_idh_distrital_atributos.csv en el
    directorio de trabajo.
"""

import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ─── 0. CONFIGURACIÓN ────────────────────────────────────────────────────────

CSV_DEFAULT = "peru_idh_distrital_atributos.csv"
OUTPUT_XLSX = "idh_distrital_limpio.xlsx"
OUTPUT_CSV  = "idh_distrital_limpio.csv"

# Colores corporativos
COLOR_HEADER   = "1D6FA8"   # azul oscuro
COLOR_SUBHEAD  = "D9EAF5"   # azul muy claro
COLOR_Q5       = "1D9E75"   # verde (quintil alto)
COLOR_Q4       = "86CBA8"
COLOR_Q3       = "FAC775"   # ámbar
COLOR_Q2       = "EF9F27"
COLOR_Q1       = "D85A30"   # rojo (quintil bajo)
COLOR_WARN     = "FFF2CC"   # amarillo advertencia
COLOR_WHITE    = "FFFFFF"

QUINTIL_COLORS = {1: COLOR_Q1, 2: COLOR_Q2, 3: COLOR_Q3, 4: COLOR_Q4, 5: COLOR_Q5}

thin = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def header_style(cell, color=COLOR_HEADER):
    cell.font = Font(bold=True, color=COLOR_WHITE, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=color, end_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN

def subheader_style(cell, color=COLOR_SUBHEAD):
    cell.font = Font(bold=True, color="1D4F73", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=color, end_color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER_THIN

def data_cell(cell, align="left"):
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal=align)
    cell.border = BORDER_THIN

def freeze_and_filter(ws, freeze="A2", filter_range=None):
    ws.freeze_panes = freeze
    if filter_range:
        ws.auto_filter.ref = filter_range

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ─── 1. CARGA Y VALIDACIÓN ───────────────────────────────────────────────────

def cargar_y_validar(ruta_csv: str) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  CARGA Y VALIDACIÓN  —  {ruta_csv}")
    print(f"{'='*60}")

    if not os.path.exists(ruta_csv):
        sys.exit(f"[ERROR] No se encontró el archivo: {ruta_csv}")

    df = pd.read_csv(
        ruta_csv,
        dtype={
            "UBIGEO"  : str,
            "COD_DPTO": str,
            "COD_PROV": str,
            "COD_DIST": str,
        },
        encoding="utf-8",
    )

    # Registros
    n = len(df)
    print(f"  Registros leídos    : {n}")
    assert n == 1874, f"Se esperaban 1,874 registros, se encontraron {n}"
    print(f"  ✓ Registros confirmados: 1,874")

    # UBIGEO
    ubigeo_len = df["UBIGEO"].str.len().value_counts()
    if set(ubigeo_len.index) != {6}:
        print(f"  [ADVERTENCIA] UBIGEO con longitud diferente a 6:\n{ubigeo_len}")
    else:
        print(f"  ✓ Todos los UBIGEO tienen exactamente 6 dígitos")

    dups = df["UBIGEO"].duplicated().sum()
    print(f"  ✓ UBIGEO duplicados: {dups}")

    nulos_idh = df["IDH"].isnull().sum()
    print(f"  ✓ IDH nulos        : {nulos_idh}")

    # Columnas requeridas
    requeridas = ["UBIGEO","IDH","NOM_DPTO","NOM_PROV","NOM_DIST",
                  "IFP_NS_MES","EVN_ANNOS","AE_ANNOS","PES_CO_POR"]
    faltantes = [c for c in requeridas if c not in df.columns]
    if faltantes:
        sys.exit(f"[ERROR] Columnas faltantes: {faltantes}")
    print(f"  ✓ Todas las columnas requeridas presentes")

    return df


# ─── 2. TRANSFORMACIÓN ───────────────────────────────────────────────────────

def transformar(df: pd.DataFrame) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  TRANSFORMACIÓN")
    print(f"{'='*60}")

    # Copia de trabajo
    out = df[["UBIGEO","NOM_DPTO","NOM_PROV","NOM_DIST",
              "IDH","IFP_NS_MES","EVN_ANNOS","AE_ANNOS","PES_CO_POR",
              "POB_HAB","IDH_RANK"]].copy()

    out = out.rename(columns={
        "IFP_NS_MES" : "ingreso_per_capita",
        "EVN_ANNOS"  : "esperanza_vida",
        "AE_ANNOS"   : "anios_educacion",
        "PES_CO_POR" : "pct_secundaria_completa",
    })

    # Quintiles: 1 = más bajo, 5 = más alto
    out["quintil_idh"] = pd.qcut(
        out["IDH"],
        q=5,
        labels=[1, 2, 3, 4, 5],
        duplicates="drop",
    ).astype(int)

    # Verificación de quintiles
    dist_q = out["quintil_idh"].value_counts().sort_index()
    print(f"  Distribución de quintiles (recalculados):")
    nombres_q = {1:"Muy bajo", 2:"Bajo", 3:"Medio", 4:"Alto", 5:"Muy alto"}
    for q, cnt in dist_q.items():
        print(f"    Quintil {q} ({nombres_q[q]:8s}): {cnt:4d} distritos")

    # Orden de columnas final
    out = out[[
        "UBIGEO","NOM_DPTO","NOM_PROV","NOM_DIST",
        "IDH","quintil_idh",
        "ingreso_per_capita","esperanza_vida","anios_educacion",
        "pct_secundaria_completa","POB_HAB","IDH_RANK",
    ]].sort_values("IDH_RANK")

    print(f"\n  IDH mín : {out['IDH'].min():.6f}")
    print(f"  IDH máx : {out['IDH'].max():.6f}")
    print(f"  IDH med : {out['IDH'].median():.6f}")
    print(f"  IDH prom: {out['IDH'].mean():.6f}")

    return out


# ─── 3. ANÁLISIS EXPLORATORIO ─────────────────────────────────────────────────

def analisis_exploratorio(df_limpio: pd.DataFrame) -> dict:
    print(f"\n{'='*60}")
    print(f"  ANÁLISIS EXPLORATORIO")
    print(f"{'='*60}")

    resultados = {}

    # 3a. Distribución por departamento
    by_dpto = df_limpio.groupby("NOM_DPTO").agg(
        n_distritos   =("IDH","count"),
        idh_promedio  =("IDH","mean"),
        idh_maximo    =("IDH","max"),
        idh_minimo    =("IDH","min"),
        rango_idh     =("IDH", lambda x: x.max() - x.min()),
        poblacion_total=("POB_HAB","sum"),
    ).round(4).reset_index().sort_values("idh_promedio", ascending=False)
    resultados["por_departamento"] = by_dpto
    print(f"\n  Distribución IDH por departamento:")
    print(by_dpto.to_string(index=False))

    # 3b. 20 distritos con IDH más bajo
    bot20 = df_limpio.nsmallest(20, "IDH")[
        ["UBIGEO","NOM_DPTO","NOM_PROV","NOM_DIST",
         "IDH","quintil_idh","ingreso_per_capita",
         "esperanza_vida","anios_educacion","POB_HAB"]
    ].copy()
    resultados["idh_mas_bajo"] = bot20
    print(f"\n  20 distritos con IDH más bajo:")
    print(bot20[["NOM_DPTO","NOM_DIST","IDH"]].to_string(index=False))

    # 3c. 20 distritos con IDH más alto
    top20 = df_limpio.nlargest(20, "IDH")[
        ["UBIGEO","NOM_DPTO","NOM_PROV","NOM_DIST",
         "IDH","quintil_idh","ingreso_per_capita",
         "esperanza_vida","anios_educacion","POB_HAB"]
    ].copy()
    resultados["idh_mas_alto"] = top20
    print(f"\n  20 distritos con IDH más alto:")
    print(top20[["NOM_DPTO","NOM_DIST","IDH"]].to_string(index=False))

    # 3d. Brecha quintil 1 vs quintil 5
    q1 = df_limpio[df_limpio["quintil_idh"] == 1]
    q5 = df_limpio[df_limpio["quintil_idh"] == 5]
    brecha_rows = []
    for col, label in [
        ("IDH",              "IDH"),
        ("ingreso_per_capita","Ingreso per cápita (S/./mes)"),
        ("esperanza_vida",   "Esperanza de vida (años)"),
        ("anios_educacion",  "Años de educación"),
        ("pct_secundaria_completa","% Secundaria completa"),
    ]:
        prom_q1 = q1[col].mean()
        prom_q5 = q5[col].mean()
        ratio = prom_q5 / prom_q1 if prom_q1 != 0 else np.nan
        brecha_rows.append({
            "Indicador"       : label,
            "Promedio Quintil 1 (más bajo)": round(prom_q1, 3),
            "Promedio Quintil 5 (más alto)": round(prom_q5, 3),
            "Ratio Q5/Q1"     : round(ratio, 2),
        })
    brecha_df = pd.DataFrame(brecha_rows)
    resultados["brecha_quintiles"] = brecha_df
    print(f"\n  Brecha Quintil 5 vs Quintil 1:")
    print(brecha_df.to_string(index=False))

    # 3e. Varianza de componentes (con coeficiente de variación)
    componentes = {
        "IDH"                    : "IDH",
        "ingreso_per_capita"     : "Ingreso per cápita",
        "anios_educacion"        : "Años de educación",
        "pct_secundaria_completa": "% Secundaria completa",
        "esperanza_vida"         : "Esperanza de vida",
    }
    var_rows = []
    for col, label in componentes.items():
        s = df_limpio[col]
        var_rows.append({
            "Componente"           : label,
            "Media"                : round(s.mean(), 3),
            "Desviación estándar"  : round(s.std(), 3),
            "Varianza"             : round(s.var(), 4),
            "Coef. variación (CV)" : round(s.std() / s.mean(), 4),
            "Mínimo"               : round(s.min(), 3),
            "Máximo"               : round(s.max(), 3),
        })
    var_df = pd.DataFrame(var_rows).sort_values("Coef. variación (CV)", ascending=False)
    resultados["varianza_componentes"] = var_df

    componente_max_cv = var_df.iloc[0]["Componente"]
    print(f"\n  Varianza de componentes (ordenado por CV):")
    print(var_df.to_string(index=False))
    print(f"\n  → Componente con mayor varianza relativa: {componente_max_cv}")

    return resultados


# ─── 4. EXPORTAR EXCEL ───────────────────────────────────────────────────────

def exportar_excel(df_limpio: pd.DataFrame, resultados: dict, ruta: str):
    print(f"\n{'='*60}")
    print(f"  EXPORTANDO EXCEL  →  {ruta}")
    print(f"{'='*60}")

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # ── HOJA 1: README ───────────────────────────────────────────────────────
    ws_readme = wb.create_sheet("README")
    ws_readme.column_dimensions["A"].width = 18
    ws_readme.column_dimensions["B"].width = 80

    readme_content = [
        ("PROYECTO",    "Geotón Perú 2026 — Índice de Vulnerabilidad Digital Compuesta (IVDC)"),
        ("ARCHIVO",     "idh_distrital_limpio.xlsx"),
        ("FUENTE",      "PNUD Perú, 2019. Unidad del Informe sobre Desarrollo Humano."),
        ("FECHA GEN.",  pd.Timestamp.today().strftime("%Y-%m-%d")),
        ("",            ""),
        ("HOJAS",       ""),
        ("  01_datos_limpios",   "1,874 distritos con IDH y componentes. Llave: UBIGEO."),
        ("  02_por_departamento","IDH promedio, rango y población por departamento."),
        ("  03_idh_mas_bajo",    "Los 20 distritos con menor IDH."),
        ("  04_idh_mas_alto",    "Los 20 distritos con mayor IDH."),
        ("  05_brecha_quintiles","Comparación Quintil 1 (más bajo) vs Quintil 5 (más alto)."),
        ("  06_varianza_componentes","Qué componente del IDH varía más entre distritos."),
        ("",            ""),
        ("CONVENCIÓN QUINTILES", ""),
        ("  quintil_idh = 1", "IDH MÁS BAJO  (≈ 0.091 – 0.28)  — Muy vulnerable"),
        ("  quintil_idh = 5", "IDH MÁS ALTO  (≈ 0.54 – 0.845)  — Menos vulnerable"),
        ("  NOTA",      "El dataset original (COD_QUINT) usaba la convención inversa (1=más alto)."),
        ("  NOTA",      "Aquí se recalcularon con pd.qcut sobre los 1,874 distritos."),
        ("",            ""),
        ("⚠ ADVERTENCIA UBIGEO", ""),
        ("",  "UBIGEO es un código de 6 dígitos con ceros iniciales (ej. '010101')."),
        ("",  "Si copias el CSV a Excel con doble clic, Excel puede convertirlo a número"),
        ("",  "y PERDER los ceros (ej. '010101' → 10101). Para evitarlo:"),
        ("",  "  Opción A) Importar el CSV con 'Datos > Desde texto/CSV' y elegir"),
        ("",  "           Formato: Texto para la columna UBIGEO."),
        ("",  "  Opción B) Usar este archivo .xlsx directamente (columna ya formateada)."),
        ("",            ""),
        ("CRUCE CON IVDC","El UBIGEO es la llave de cruce con el Índice de Vulnerabilidad"),
        ("",            "Digital Compuesta (IVDC). Los distritos con quintil_idh=1 o 2"),
        ("",            "deberían mostrar también alta vulnerabilidad digital."),
    ]

    ws_readme["A1"] = "IDH Distrital Perú — Documentación"
    ws_readme["A1"].font = Font(bold=True, size=13, name="Arial", color=COLOR_HEADER)
    ws_readme.row_dimensions[1].height = 22

    for i, (key, val) in enumerate(readme_content, start=3):
        cell_k = ws_readme.cell(row=i, column=1, value=key)
        cell_v = ws_readme.cell(row=i, column=2, value=val)
        cell_k.font = Font(bold=bool(key and not key.startswith(" ")), name="Arial", size=9)
        cell_v.font = Font(name="Arial", size=9)
        if key.startswith("⚠"):
            cell_k.font = Font(bold=True, name="Arial", size=9, color="FF6600")
            cell_k.fill = PatternFill("solid", start_color=COLOR_WARN, end_color=COLOR_WARN)
            cell_v.fill = PatternFill("solid", start_color=COLOR_WARN, end_color=COLOR_WARN)

    print("  ✓ Hoja README")

    # ── HOJA 2: datos limpios ─────────────────────────────────────────────────
    ws_data = wb.create_sheet("01_datos_limpios")

    col_headers = {
        "UBIGEO"                 : "UBIGEO\n(6 dígitos)",
        "NOM_DPTO"               : "Departamento",
        "NOM_PROV"               : "Provincia",
        "NOM_DIST"               : "Distrito",
        "IDH"                    : "IDH",
        "quintil_idh"            : "Quintil IDH\n(1=bajo·5=alto)",
        "ingreso_per_capita"     : "Ingreso per cápita\n(S/./mes)",
        "esperanza_vida"         : "Esperanza de vida\n(años)",
        "anios_educacion"        : "Años de\neducación",
        "pct_secundaria_completa": "% Secundaria\ncompleta",
        "POB_HAB"                : "Población\n(hab.)",
        "IDH_RANK"               : "Ranking\nIDH",
    }
    col_widths = [12, 16, 18, 22, 8, 14, 16, 14, 12, 14, 13, 10]
    col_aligns = ["center","left","left","left","center","center",
                  "right","right","right","right","right","center"]
    col_formats = [None, None, None, None,
                   "0.0000", None, "#,##0.0", "0.00", "0.00", "0.00",
                   "#,##0", "#,##0"]

    headers = list(col_headers.values())
    ws_data.row_dimensions[1].height = 30
    for c_idx, h in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=c_idx, value=h)
        header_style(cell)
        ws_data.column_dimensions[get_column_letter(c_idx)].width = col_widths[c_idx-1]

    for r_idx, row in enumerate(df_limpio.itertuples(index=False), start=2):
        for c_idx, col in enumerate(df_limpio.columns, start=1):
            val = getattr(row, col)
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            align = col_aligns[c_idx-1]
            fmt   = col_formats[c_idx-1]
            data_cell(cell, align=align)
            if fmt:
                cell.number_format = fmt
            # UBIGEO: forzar texto
            if col == "UBIGEO":
                cell.number_format = "@"
            # Color por quintil
            if col == "quintil_idh" and isinstance(val, (int, np.integer)):
                q_color = QUINTIL_COLORS.get(int(val), COLOR_WHITE)
                cell.fill = PatternFill("solid", start_color=q_color, end_color=q_color)
                cell.font = Font(bold=True, name="Arial", size=9,
                                 color=COLOR_WHITE if int(val) in (1,5) else "000000")

    freeze_and_filter(ws_data, freeze="E2",
                      filter_range=f"A1:{get_column_letter(len(headers))}1")
    print("  ✓ Hoja 01_datos_limpios")

    # ── HOJA 3: por departamento ──────────────────────────────────────────────
    ws_dpto = wb.create_sheet("02_por_departamento")
    dpto_df = resultados["por_departamento"]

    dpto_headers = {
        "NOM_DPTO"       : "Departamento",
        "n_distritos"    : "N° Distritos",
        "idh_promedio"   : "IDH Promedio",
        "idh_maximo"     : "IDH Máximo",
        "idh_minimo"     : "IDH Mínimo",
        "rango_idh"      : "Rango IDH\n(max - min)",
        "poblacion_total": "Población\nTotal",
    }
    ws_dpto.row_dimensions[1].height = 30
    for c_idx, h in enumerate(dpto_headers.values(), start=1):
        cell = ws_dpto.cell(row=1, column=c_idx, value=h)
        header_style(cell)
    ws_dpto.column_dimensions["A"].width = 18
    for ci in range(2, len(dpto_headers)+1):
        ws_dpto.column_dimensions[get_column_letter(ci)].width = 13

    for r_idx, row in enumerate(dpto_df.itertuples(index=False), start=2):
        for c_idx, col in enumerate(dpto_df.columns, start=1):
            val = getattr(row, col)
            cell = ws_dpto.cell(row=r_idx, column=c_idx, value=val)
            data_cell(cell, align="right" if c_idx > 1 else "left")
            if c_idx in (3,4,5,6):
                cell.number_format = "0.0000"
            elif c_idx == 7:
                cell.number_format = "#,##0"

    freeze_and_filter(ws_dpto, freeze="B2", filter_range=f"A1:G1")
    print("  ✓ Hoja 02_por_departamento")

    # ── HOJA 4 y 5: top/bottom 20 ────────────────────────────────────────────
    for sheet_name, df_sub, titulo in [
        ("03_idh_mas_bajo", resultados["idh_mas_bajo"], "20 distritos con IDH más BAJO"),
        ("04_idh_mas_alto", resultados["idh_mas_alto"], "20 distritos con IDH más ALTO"),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, name="Arial", size=11, color=COLOR_HEADER)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 28

        sub_headers = ["UBIGEO","Departamento","Provincia","Distrito",
                       "IDH","Quintil IDH","Ingreso per cápita","Esp. vida","Años educ.","Población"]
        sub_widths  = [12, 16, 18, 22, 8, 10, 16, 10, 10, 13]
        sub_formats = [None,None,None,None,"0.0000",None,"#,##0.0","0.00","0.00","#,##0"]

        for c_idx, (h, w) in enumerate(zip(sub_headers, sub_widths), start=1):
            cell = ws.cell(row=2, column=c_idx, value=h)
            header_style(cell)
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        for r_idx, row in enumerate(df_sub.itertuples(index=False), start=3):
            for c_idx, col in enumerate(df_sub.columns, start=1):
                val = getattr(row, col)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                data_cell(cell, align="right" if c_idx > 4 else "left")
                fmt = sub_formats[c_idx-1]
                if fmt:
                    cell.number_format = fmt
                if col == "UBIGEO":
                    cell.number_format = "@"
                # Fila alternada
                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="F2F7FB", end_color="F2F7FB")

        ws.freeze_panes = "E3"
        print(f"  ✓ Hoja {sheet_name}")

    # ── HOJA 6: brecha quintiles ──────────────────────────────────────────────
    ws_bq = wb.create_sheet("05_brecha_quintiles")
    bq_df = resultados["brecha_quintiles"]
    ws_bq["A1"] = "Brecha entre Quintil 1 (IDH más bajo) y Quintil 5 (IDH más alto)"
    ws_bq["A1"].font = Font(bold=True, name="Arial", size=11, color=COLOR_HEADER)
    ws_bq.row_dimensions[1].height = 18
    ws_bq.row_dimensions[2].height = 28
    ws_bq.column_dimensions["A"].width = 28
    for ci in range(2, 5):
        ws_bq.column_dimensions[get_column_letter(ci)].width = 24

    for c_idx, h in enumerate(bq_df.columns, start=1):
        cell = ws_bq.cell(row=2, column=c_idx, value=h)
        if "Quintil 1" in h:
            header_style(cell, color=COLOR_Q1)
        elif "Quintil 5" in h:
            header_style(cell, color=COLOR_Q5)
        else:
            header_style(cell)

    for r_idx, row_vals in enumerate(bq_df.values.tolist(), start=3):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_bq.cell(row=r_idx, column=c_idx, value=val)
            data_cell(cell, align="right" if c_idx > 1 else "left")
            if c_idx == 2:
                cell.fill = PatternFill("solid", start_color="FDECEA", end_color="FDECEA")
            elif c_idx == 3:
                cell.fill = PatternFill("solid", start_color="E8F5EE", end_color="E8F5EE")

    print("  ✓ Hoja 05_brecha_quintiles")

    # ── HOJA 7: varianza de componentes ──────────────────────────────────────
    ws_var = wb.create_sheet("06_varianza_componentes")
    var_df = resultados["varianza_componentes"]
    ws_var["A1"] = "Varianza de componentes del IDH — ¿Qué varía más entre distritos?"
    ws_var["A1"].font = Font(bold=True, name="Arial", size=11, color=COLOR_HEADER)
    ws_var.row_dimensions[1].height = 18
    ws_var.row_dimensions[2].height = 28

    var_widths = [26, 12, 20, 12, 22, 12, 12]
    for c_idx, (h, w) in enumerate(zip(var_df.columns, var_widths), start=1):
        cell = ws_var.cell(row=2, column=c_idx, value=h)
        header_style(cell)
        ws_var.column_dimensions[get_column_letter(c_idx)].width = w

    for r_idx, row_vals in enumerate(var_df.values.tolist(), start=3):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_var.cell(row=r_idx, column=c_idx, value=val)
            data_cell(cell, align="right" if c_idx > 1 else "left")
        if r_idx == 3:  # Mayor CV
            for c_idx in range(1, len(var_df.columns)+1):
                ws_var.cell(row=3, column=c_idx).fill = PatternFill(
                    "solid", start_color="FFF0E0", end_color="FFF0E0"
                )
            ws_var.cell(row=3, column=1).font = Font(bold=True, name="Arial", size=9)

    nota = ws_var.cell(row=len(var_df)+4, column=1,
                       value="Nota: CV = Coef. de Variación = std / media. "
                             "Mayor CV indica mayor dispersión relativa entre distritos.")
    nota.font = Font(italic=True, name="Arial", size=9, color="666666")

    print("  ✓ Hoja 06_varianza_componentes")

    wb.save(ruta)
    print(f"\n  ✅ Excel guardado: {ruta}")


# ─── 5. EXPORTAR CSV LIMPIO ──────────────────────────────────────────────────

def exportar_csv(df_limpio: pd.DataFrame, ruta: str):
    df_limpio.to_csv(ruta, index=False, encoding="utf-8-sig")
    print(f"  ✅ CSV guardado   : {ruta}")
    print(f"     (utf-8-sig para compatibilidad con Excel en Windows)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    ruta_csv = sys.argv[1] if len(sys.argv) > 1 else CSV_DEFAULT

    df_raw    = cargar_y_validar(ruta_csv)
    df_limpio = transformar(df_raw)
    resultados = analisis_exploratorio(df_limpio)
    exportar_csv(df_limpio, OUTPUT_CSV)
    exportar_excel(df_limpio, resultados, OUTPUT_XLSX)

    print(f"\n{'='*60}")
    print(f"  ✅ PROCESO COMPLETADO")
    print(f"     {OUTPUT_CSV}")
    print(f"     {OUTPUT_XLSX}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
